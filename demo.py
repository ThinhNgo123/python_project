import json
import openpyxl



class ExcelManipulation:
    def excelExport(self, data, path):
        """ this function export the user information to a file.xlsx

         Args:
            data: array of dictionary of user information
            path: the path to save file.xlsx
        """

        # initialize excel object
        excelFile = openpyxl.Workbook()
        excelFileActive = excelFile.active

        # size of data 
        dataLen = len(data)

        if dataLen == 0:
            print("Can't export!")
        else:
            fieldsOfData = []
            lenFieldOfData = 0

            # array name fieldsOfData include fields 
            # and write in the cells in the first line in the excel file 
            for field in data[0]:
                fieldsOfData.append(field)
                excelFileActive.cell(row = 1, column = lenFieldOfData + 1, value = field)
                lenFieldOfData += 1

            # write data into the next cells in the excel file
            for i in range(dataLen):
                for j in range(lenFieldOfData):
                    excelFileActive.cell(row = i + 2, column = j + 1, value = data[i][fieldsOfData[j]])
            
            # save file excel
            excelFile.save(path)

    def excelImport(path, sheet, data):
        """ This function import from file.xlsx to data contains user information

        Args:
            path: the path to file.xlsx
            sheet: sheet name in file.xlsx
            data: array of dictionary of user information
        """

        excelFile = openpyxl.load_workbook(path)
        sheetFile = excelFile[sheet]

        fields = []

        data.clear()
        
        for row in sheetFile:
            for cell in row:
                fields.append(cell.value)
            break
        
        check = 0
        
        for row in sheetFile:
            if check == 0:
                check = 1
                continue
            index = 0
            dict = {}
            for cell in row:
                dict[fields[index]] = str(cell.value)
                index += 1
            data.append(dict)
        
        excelFile.close()

class JsonManipulation:
    def saveFile(data : any, path : any):
        """

        Args:
            data (_type_): _description_
            path (_type_): _description_
        """
        dataLen = len(data)

        index = 1

        dict = {}

        with open(path, "w") as jsonFile:
            for i in range(dataLen):
                dict[str(index)] = data[i]
                index += 1

            data_string = json.dumps(dict, indent = 4)
            jsonFile.write(data_string)
            jsonFile.close()

    def readFile(path, data):
        """_summary_

        Args:
            path (_type_): _description_
            data (_type_): _description_
        """
        with open(path, "r") as jsonFile:
            dict = json.load(jsonFile)

            for value in dict.values():
                data.append(value)

            jsonFile.close()

class UserManager:
    def userUpdate(data, pos, field, value):
        """

        Args:
            data (_type_): _description_
            pos (_type_): _description_
            field (_type_): _description_
            value (_type_): _description_
        """

        data[pos - 1][field] = value
        print("New information has been updated!")

    def userDelete(data, pos):
        """_summary_

        Args:
            data (_type_): _description_
            pos (_type_): _description_
        """

        data.pop(pos - 1)
        print("User has been deleted!")

def main():
    EXPORT = 1
    IMPORT = 2
    UPDATE = 3
    DELETE = 4
    SAVE = 5
    LOAD = 6
    EXIT = 0

    path = "data1.json"

    data = []

    JsonManipulation.readFile(path, data)

    while(True):
        print("------------------")
        print("1. Export to file excel")
        print("2. Import from file excel")
        print("3. Update user information")
        print("4. Delete user")
        print("5. Save file json")
        print("6. Load file json")
        print("0. Exit")

        select = int(input("Choose: "))

        if(select == EXPORT):
            path1 = input("path: ")
            ExcelManipulation.excelExport(data, path1)
            print("Export success to excel file!")

        elif(select == IMPORT):
            path1 = input("path: ")
            sheet = input("sheet: ")
            ExcelManipulation.excelImport(path1, sheet, data)
            JsonManipulation.saveFile(data, path)
            print("Import success!")

        elif(select == UPDATE):
            pos = int(input("User's sequence number: "))
            field = input("Field need change ('userName', 'password', 'age', 'sex'): ")
            value = input("New information: ")
            UserManager.userUpdate(data, pos, field, value)
            JsonManipulation.saveFile(data, path)

        elif(select == DELETE):
            pos = int(input("User's sequence number: "))
            UserManager.userDelete(data, pos)
            JsonManipulation.saveFile(data, path)

        elif(select == SAVE):
            path1 = input("path: ")
            JsonManipulation.saveFile(data, path1)
            print("Save file success!")

        elif(select == LOAD):
            path1 = input("path: ")
            data1 = []
            JsonManipulation.readFile(path1, data1)
            print("Load file success!")

        elif(select == EXIT):
            break

        else:
            select = int(input("Retype: "))

if __name__ == "__main__":
    main()