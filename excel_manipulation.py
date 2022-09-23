import openpyxl

class ExcelManipulation:
    # COLUMN_NUMBER = 5
    SAVE_FILE_SUCCESS = "Save file success"
    SAVE_FILE_ERROR = "Save file error"
    READ_FILE_ERROR = "Read file error"
    READ_FILE_SUCCESS = "Read file success"
    DELETE_SUCCESS = "Delete success"
    DELETE_ERROR = "Delete error"
    # DEFAULT_END_ROW = 1
    DEFAULT_MAX_USER = 100
    DEFAULT_FILED = ["Index", "Station name", "IP", "User", "Password"]

    def __init__(self, path = "data_excel.xlsx", sheet = "Sheet", maxUser = DEFAULT_MAX_USER):
        self.path = path
        self.sheet = sheet
        self.maxUser = maxUser
        # self.endRowFileExcel = self.DEFAULT_END_ROW
        # self.data = []
        # self.excelImport()

    # def openExcelFile(self):
    #     self.excelFile = openpyxl.load_workbook(self.path)
    #     self.sheetFile = self.excelFile[self.sheet]

    def createExcelFile(self):
        '''
        This function used to create file excel new
        '''
        excelFile = openpyxl.Workbook()
        excelFileActive = excelFile.active

        for col, filed in enumerate(self.DEFAULT_FILED):
            excelFileActive.cell(row = 1, column = col + 1, value = filed)

        for index in range(self.maxUser):
            excelFileActive.cell(row = index + 2, column = 1, value = index)

        excelFile.save(self.path)

    def getUserFromExcel(self, index):
        ''' this function export the user information to a file.xlsx
        - param: index
        - @return: idx, stationName, ip, user, password
        '''
        try:
            excelFile = openpyxl.load_workbook(self.path)
            sheetFile = excelFile[self.sheet]

            index = int(index)

            # get information user from excel file
            idx         = sheetFile.cell(row = index + 2, column = 1).value
            stationName = sheetFile.cell(row = index + 2, column = 2).value
            if stationName == None:
                return "", "", "", "", ""
            ip          = sheetFile.cell(row = index + 2, column = 3).value
            user        = sheetFile.cell(row = index + 2, column = 4).value
            password    = sheetFile.cell(row = index + 2, column = 5).value
            
            # close excel file
            excelFile.close()

            return idx, stationName, ip, user, password
        except:
            return "", "", "", "", ""

    def updateUserToExcel(self, index, stationName, ip, user, password):
        ''' this function used to save information user after replacing
        - param: index
        '''

        try:
            # open file excel
            excelFile = openpyxl.load_workbook(self.path)
            sheetFile = excelFile[self.sheet]

            for col, filed in enumerate([index, stationName, ip, user, password], 1):
                sheetFile.cell(row = int(index) + 2, column = col, value = filed)

            excelFile.save(self.path)

            return self.SAVE_FILE_SUCCESS
        except:
            # create file excel
            self.createExcelFile()

            self.updateUserToExcel(index, stationName, ip, user, password)

    def addMaxUser(self, maxUser):
        if maxUser > self.maxUser:
            try:
                excelFile = openpyxl.load_workbook(self.path)
                sheetFile = excelFile[self.sheet]

                for row in range(self.maxUser, maxUser):
                    sheetFile.cell(row = row + 2, column = 1, value = row)

                excelFile.save(self.path)
                self.maxUser = maxUser
            except:
                self.maxUser = maxUser
                self.createExcelFile()


    def excelImport(self):
        ''' This function import from file.xlsx and return array of dictionary
        of user information

        - @return:
        '''
        try:
            excelFile = openpyxl.load_workbook(self.path)
            sheetFile = excelFile[self.sheet]
        except:
            self.createExcelFile()

            excelFile = openpyxl.load_workbook(self.path)
            sheetFile = excelFile[self.sheet]
        finally:
            fields = []
            data = []
            
            for row in sheetFile:
                for cell in row:
                    fields.append(cell.value)
                break
            
            check = 0
            
            for row in sheetFile:
                if check == 0:
                    check = 1
                    continue
                index = 0
                dict = {}
                for cell in row:
                    dict[fields[index]] = str(cell.value)
                    index += 1
                data.append(dict)
            
            excelFile.close()
            
            return data

    def deleteFromRow(self, index):
        if 0 <= index <= self.maxUser - 1:
            try:
                # count = int(index)
                excelFile = openpyxl.load_workbook(self.path)
                sheetFile = excelFile[self.sheet]

                for col in range(2, len(self.DEFAULT_FILED) + 1):
                    sheetFile.cell(row = int(index) + 2, column = col, value = "") 

                excelFile.save(self.path)
                return self.DELETE_SUCCESS
            except:
                return self.DELETE_ERROR
        return self.DELETE_ERROR
